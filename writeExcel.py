from openpyxl import load_workbook

#https://openpyxl.readthedocs.io/en/stable/usage.html

wb = load_workbook(filename = 'Projektübersicht.xlsx')
sheet_ranges = wb['Tabelle1']
print(sheet_ranges['A2'].value)
new_sheet = wb.create_sheet(title="Data")
for row in range(1,10):
    new_sheet.cell(row=row,column=15,value=row)

wb.save(filename = 'Projektübersicht.xlsx')


